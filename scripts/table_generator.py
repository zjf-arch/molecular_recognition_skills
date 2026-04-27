"""
表格生成模块
用于将识别结果输出为表格（Excel/CSV/Markdown）
"""

import pandas as pd
from pathlib import Path
from typing import List, Dict, Any
import base64
import logging
from datetime import datetime

logger = logging.getLogger(__name__)


class TableGenerator:
    """表格生成器"""

    def __init__(self, output_dir: str):
        """
        初始化表格生成器

        Args:
            output_dir: 输出目录
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def create_result_table(
        self,
        results: List[Dict[str, Any]],
        original_patent_name: str,
        include_images: bool = True,
        format: str = 'xlsx'
    ) -> str:
        """
        创建结果表格

        Args:
            results: 识别结果列表，每个元素包含：
                - index: 序号
                - image_path: 分子结构图路径
                - smiles: SMILES 字符串
                - success: 是否成功
            original_patent_name: 原始专利文件名
            include_images: 是否在表格中包含图片
            format: 输出格式 (xlsx, csv, md)

        Returns:
            生成的表格文件路径
        """
        # 准备数据
        data = []

        for result in results:
            row = {
                '序号': result.get('index', 0),
                '分子结构图': result.get('image_path', ''),
                'SMILES': result.get('smiles', '') if result.get('success') else '识别失败',
                '识别状态': '成功' if result.get('success') else '失败',
                '原文献名称': original_patent_name
            }

            # 如果需要包含图片且识别成功
            if include_images and result.get('success') and result.get('image_path'):
                try:
                    image_path = result['image_path']
                    if Path(image_path).exists():
                        # 对于 Excel，我们可以嵌入图片（需要使用 openpyxl）
                        # 这里先存储路径，后面处理
                        row['分子结构图路径'] = image_path
                except Exception as e:
                    logger.warning(f"处理图片时出错: {e}")

            data.append(row)

        # 创建 DataFrame
        df = pd.DataFrame(data)

        # 生成文件名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        base_name = Path(original_patent_name).stem
        output_filename = f"{base_name}_分子结构_{timestamp}"

        # 根据格式输出
        if format == 'xlsx':
            output_path = self._save_as_excel(df, output_filename, include_images, data)
        elif format == 'csv':
            output_path = self._save_as_csv(df, output_filename)
        elif format == 'md':
            output_path = self._save_as_markdown(df, output_filename)
        else:
            logger.warning(f"不支持的格式: {format}，使用默认 xlsx")
            output_path = self._save_as_excel(df, output_filename, include_images, data)

        logger.info(f"表格已生成: {output_path}")
        return output_path

    def _save_as_excel(
        self,
        df: pd.DataFrame,
        filename: str,
        include_images: bool,
        original_data: List[Dict]
    ) -> str:
        """
        保存为 Excel 格式

        Args:
            df: 数据框
            filename: 文件名（不含扩展名）
            include_images: 是否包含图片
            original_data: 原始数据（用于获取图片路径）

        Returns:
            文件路径
        """
        output_path = self.output_dir / f"{filename}.xlsx"

        if include_images:
            # 使用 openpyxl 嵌入图片
            from openpyxl import Workbook
            from openpyxl.drawing.image import Image as XLImage
            from openpyxl.utils.dataframe import dataframe_to_rows

            wb = Workbook()
            ws = wb.active
            ws.title = "分子结构识别结果"

            # 写入表头
            headers = ['序号', '分子结构图', 'SMILES', '识别状态', '原文献名称']
            ws.append(headers)

            # 设置列宽
            ws.column_dimensions['A'].width = 8   # 序号
            ws.column_dimensions['B'].width = 30  # 分子结构图
            ws.column_dimensions['C'].width = 50  # SMILES
            ws.column_dimensions['D'].width = 12  # 识别状态
            ws.column_dimensions['E'].width = 30  # 原文献名称

            # 写入数据
            for idx, row in enumerate(original_data, start=2):
                ws.cell(row=idx, column=1, value=row['序号'])
                ws.cell(row=idx, column=3, value=row['SMILES'])
                ws.cell(row=idx, column=4, value=row['识别状态'])
                ws.cell(row=idx, column=5, value=row['原文献名称'])

                # 插入图片
                image_path = row.get('分子结构图路径') or row.get('分子结构图')
                if image_path and Path(image_path).exists():
                    try:
                        img = XLImage(image_path)
                        # 调整图片大小
                        img.width = 150
                        img.height = 100
                        # 添加图片到单元格
                        ws.add_image(img, f'B{idx}')
                        # 调整行高
                        ws.row_dimensions[idx].height = 80
                    except Exception as e:
                        logger.warning(f"插入图片失败 {image_path}: {e}")
                        ws.cell(row=idx, column=2, value=image_path)
                else:
                    ws.cell(row=idx, column=2, value=image_path)

            wb.save(output_path)
        else:
            # 不包含图片，直接保存
            df.to_excel(output_path, index=False, engine='openpyxl')

        return str(output_path)

    def _save_as_csv(self, df: pd.DataFrame, filename: str) -> str:
        """保存为 CSV 格式"""
        output_path = self.output_dir / f"{filename}.csv"
        df.to_csv(output_path, index=False, encoding='utf-8-sig')
        return str(output_path)

    def _save_as_markdown(self, df: pd.DataFrame, filename: str) -> str:
        """保存为 Markdown 格式"""
        output_path = self.output_dir / f"{filename}.md"

        # 转换为 Markdown 表格
        md_content = "# 分子结构识别结果\n\n"
        md_content += df.to_markdown(index=False)

        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(md_content)

        return str(output_path)


def create_summary_table(results_list: List[Dict[str, Any]], output_dir: str) -> str:
    """
    创建汇总表格（多个专利的结果）

    Args:
        results_list: 多个专利的结果列表
        output_dir: 输出目录

    Returns:
        汇总表格路径
    """
    generator = TableGenerator(output_dir)

    # 合并所有结果
    all_results = []
    for patent_results in results_list:
        all_results.extend(patent_results['results'])

    # 生成汇总表格
    return generator.create_result_table(
        all_results,
        original_patent_name="汇总结果",
        include_images=False,
        format='xlsx'
    )


if __name__ == "__main__":
    # 测试代码
    import logging

    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )

    # 模拟测试数据
    test_results = [
        {
            'index': 1,
            'image_path': 'test_image_1.png',
            'smiles': 'CCO',
            'success': True
        },
        {
            'index': 2,
            'image_path': 'test_image_2.png',
            'smiles': 'CC(=O)O',
            'success': True
        },
        {
            'index': 3,
            'image_path': 'test_image_3.png',
            'smiles': '',
            'success': False
        }
    ]

    generator = TableGenerator("D:\\skills\\results")
    output_path = generator.create_result_table(
        test_results,
        "test_patent.pdf",
        include_images=False,
        format='xlsx'
    )

    print(f"测试表格已生成: {output_path}")