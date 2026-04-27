#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
简化集成处理流程：分类 → 过滤 → 识别 → Ce替换 → 单表格输出
使用方式：
  python simplified_integrated_processor.py --pdf 专利.pdf --recognizer alchemist
  python simplified_integrated_processor.py --pdf 专利.pdf --recognizer decimer
"""

import sys
import os
from pathlib import Path
import argparse
import pandas as pd
from datetime import datetime
from typing import List, Dict
import cv2
import numpy as np

PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from scripts.logger import logger

# 尝试导入RDKit用于Ce替换
try:
    from rdkit import Chem
    from rdkit import RDLogger
    RDLogger.DisableLog('rdApp.*')
    RDKIT_AVAILABLE = True
except ImportError:
    RDKIT_AVAILABLE = False
    logger.warning("RDKit未安装，Ce替换功能将不可用")


class SMILESPostProcessor:
    """SMILES后处理器 - Ce/Ge替换为异丙烯基"""

    @staticmethod
    def replace_ce_with_allyl(smiles: str) -> str:
        """
        将 SMILES 中的 'Ce' 或 '[Ge]' 替换为异丙烯基 (isopropenyl group: C=C(C))

        Args:
            smiles: 包含Ce或[Ge]的SMILES字符串

        Returns:
            替换后的SMILES字符串
        """
        if not RDKIT_AVAILABLE:
            logger.warning("RDKit未安装，无法进行Ce/Ge替换")
            return smiles

        # 检查是否包含Ce或Ge标记
        has_ce = 'Ce' in smiles
        has_ge = '[Ge]' in smiles

        if not (has_ce or has_ge):
            return smiles  # 如果没有 Ce 或 Ge，直接返回原 SMILES

        try:
            # 将 Ce 或 [Ge] 替换为连接点 *
            smiles_with_star = smiles.replace('Ce', '*').replace('[Ge]', '*')

            # 解析含连接点的分子
            mol = Chem.MolFromSmiles(smiles_with_star)
            if mol is None:
                logger.warning(f"无法解析SMILES: {smiles_with_star}")
                return smiles

            # 创建可编辑分子
            rwmol = Chem.RWMol(mol)

            # 找到所有连接点 * 的索引
            star_indices = []
            for atom in rwmol.GetAtoms():
                if atom.GetSymbol() == '*':
                    star_indices.append(atom.GetIdx())

            if not star_indices:
                return Chem.MolToSmiles(mol)  # 没有连接点，直接返回

            # 异丙烯基：CH2=C(CH3)-，连接点在第一个C
            allyl_smiles = "C=C(C)"
            allyl_mol = Chem.MolFromSmiles(allyl_smiles)
            if allyl_mol is None:
                logger.error("无法构建异丙烯基片段")
                return smiles

            # 从右到左处理连接点（避免索引偏移）
            for star_idx in sorted(star_indices, reverse=True):
                # 获取连接点的邻居
                neighbors = rwmol.GetAtomWithIdx(star_idx).GetNeighbors()
                if len(neighbors) != 1:
                    logger.warning(f"连接点 * 必须只连接一个原子，但在 {smiles} 中有 {len(neighbors)} 个")
                    return smiles

                neighbor_idx = neighbors[0].GetIdx()

                # 删除连接点 *
                rwmol.RemoveAtom(star_idx)

                # 调整 neighbor_idx
                if star_idx < neighbor_idx:
                    neighbor_idx -= 1

                # 添加异丙烯基的所有原子
                atom_map = {}
                for atom in allyl_mol.GetAtoms():
                    new_atom = Chem.Atom(atom.GetSymbol())
                    new_idx = rwmol.AddAtom(new_atom)
                    atom_map[atom.GetIdx()] = new_idx

                # 添加异丙烯基内部的键
                for bond in allyl_mol.GetBonds():
                    a1 = atom_map[bond.GetBeginAtomIdx()]
                    a2 = atom_map[bond.GetEndAtomIdx()]
                    rwmol.AddBond(a1, a2, bond.GetBondType())

                # 将原分子的 neighbor 与异丙烯基的连接点连接
                allyl_connection_idx = atom_map[1]
                rwmol.AddBond(neighbor_idx, allyl_connection_idx, Chem.BondType.SINGLE)

            # 生成新的 SMILES（使用Kekulize避免自动芳香化）
            new_mol = rwmol.GetMol()
            # Kekulize：保持双键形式，不自动芳香化环结构
            Chem.Kekulize(new_mol, clearAromaticFlags=True)
            new_smiles = Chem.MolToSmiles(new_mol, isomericSmiles=True, kekuleSmiles=True)
            return new_smiles

        except Exception as e:
            logger.error(f"Ce/Ge替换失败: {e}")
            return smiles


class SMILESRecognizer:
    """SMILES识别器 - 支持多种识别引擎"""

    def __init__(self, recognizer_type: str = 'alchemist'):
        """
        初始化识别器

        Args:
            recognizer_type: 识别类型 ('alchemist' 或 'decimer')
        """
        self.recognizer_type = recognizer_type.lower()

        if self.recognizer_type == 'alchemist':
            logger.info("使用 alchemist API 识别")
            from scripts.alchemist_api import AlchemistAPI
            self.recognizer = AlchemistAPI(
                api_url="https://api-ocsr.alchemist.iresearch.net.cn/ocsr/",
                headers={"Content-Type": "application/json"}
            )

        elif self.recognizer_type == 'decimer':
            logger.info("使用 DECIMER Transformer 识别")
            try:
                # 先添加路径到sys.path
                decimer_path = PROJECT_ROOT / "DECIMER-Image_Transformer"
                if str(decimer_path) not in sys.path:
                    sys.path.insert(0, str(decimer_path))
                from DECIMER import predict_SMILES
                self.predict_SMILES = predict_SMILES  # 保存为实例变量
                logger.info("✅ DECIMER Transformer 加载成功")
            except Exception as e:
                logger.error(f"DECIMER Transformer 加载失败: {e}")
                raise
        else:
            raise ValueError(f"不支持的识别类型: {recognizer_type}，请选择 'alchemist' 或 'decimer'")

    def recognize(self, image_path: str) -> Dict:
        """
        识别分子结构图生成SMILES

        Args:
            image_path: 图片路径

        Returns:
            识别结果字典
        """
        try:
            if self.recognizer_type == 'alchemist':
                result = self.recognizer.recognize_molecule(image_path)
                result['recognizer'] = 'alchemist'
                return result

            elif self.recognizer_type == 'decimer':
                smiles = self.predict_SMILES(image_path)

                return {
                    'success': True,
                    'smiles': smiles,
                    'confidence': 1.0,
                    'recognizer': 'decimer',
                    'image_path': image_path
                }

        except Exception as e:
            logger.error(f"识别失败 ({self.recognizer_type}): {e}")
            return {
                'success': False,
                'error': str(e),
                'recognizer': self.recognizer_type
            }


class SimplifiedIntegratedProcessor:
    """简化集成处理器 - 单一流程，单一输出"""

    def __init__(self, recognizer_type: str = 'alchemist'):
        """
        初始化处理器

        Args:
            recognizer_type: 识别类型 ('alchemist' 或 'decimer')
        """
        logger.info("=" * 80)
        logger.info("简化集成专利处理系统")
        logger.info(f"识别引擎: {recognizer_type.upper()}")
        logger.info("=" * 80)

        # 1. DECIMER提取器（带页码溯源）
        logger.info("加载 DECIMER Segmentation（带页码溯源）...")
        from scripts.decimer_processor_with_page_tracking import DECIMERProcessorWithPageTracking
        # 初始化时不设置输出目录，处理时会自动在output/pdf_name下保存
        self.decimer = DECIMERProcessorWithPageTracking()

        # 2. ResNet18分类器
        logger.info("加载 ResNet18 分类器...")
        from scripts.molecule_classifier import MoleculeImageClassifier
        self.classifier = MoleculeImageClassifier()

        # 3. SMILES识别器
        logger.info(f"加载 {recognizer_type.upper()} 识别器...")
        self.recognizer = SMILESRecognizer(recognizer_type)

        logger.info("✅ 所有组件加载完成\n")

    def process_pdf(
        self,
        pdf_path: str,
        output_dir: str = None
    ) -> Dict:
        """
        处理单个PDF文件

        Args:
            pdf_path: PDF文件路径
            output_dir: 输出目录

        Returns:
            处理结果字典
        """
        pdf_path = Path(pdf_path)
        if not pdf_path.exists():
            raise FileNotFoundError(f"PDF文件不存在: {pdf_path}")

        # 输出目录 - 缩短路径避免Windows路径长度限制
        if output_dir is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            # 截断文件名避免路径过长（最多保留前30字符）
            stem_short = pdf_path.stem[:30] if len(pdf_path.stem) > 30 else pdf_path.stem
            output_dir = PROJECT_ROOT / "output" / f"{stem_short}_{timestamp}"
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        logger.info("=" * 80)
        logger.info(f"处理专利: {pdf_path.name}")
        logger.info("=" * 80)
        logger.info(f"输出目录: {output_dir}")

        # ==================== 步骤1: DECIMER提取图片 ====================
        logger.info("\n【步骤1】DECIMER 提取化学结构...")

        # 设置DECIMER输出目录 - 使用临时文件夹，稍后会删除
        temp_output = output_dir / "temp_images"
        self.decimer.output_dir = temp_output

        structures = self.decimer.extract_structures_from_pdf(
            str(pdf_path),
            expand=True,
            save_images=True
        )

        if not structures:
            logger.warning("DECIMER 未提取到任何结构")
            return {
                'success': False,
                'error': '未提取到化学结构',
                'pdf': str(pdf_path)
            }

        logger.info(f"✅ DECIMER 提取: {len(structures)} 张图片")

        # 提取图片路径和页码信息
        image_data = []
        for s in structures:
            image_data.append({
                'path': s['image_path'],
                'page_number': s.get('page_number', '未知')  # 页码信息
            })

        # ==================== 步骤2: ResNet18分类并分别保存 ====================
        logger.info("\n【步骤2】ResNet18 分类并分别保存...")

        # 创建三个分类文件夹（与临时文件夹分离）
        small_molecule_dir = output_dir / "images" / "小分子"
        polymer_dir = output_dir / "images" / "聚合物"
        discarded_dir = output_dir / "images" / "废弃"

        small_molecule_dir.mkdir(parents=True, exist_ok=True)
        polymer_dir.mkdir(parents=True, exist_ok=True)
        discarded_dir.mkdir(parents=True, exist_ok=True)

        valid_images = []  # 保留的图片（小分子 + 聚合物）
        classification_stats = {
            '小分子': 0,
            '聚合物': 0,
            '废弃': 0
        }

        # 用于记录每种类型的计数
        type_counters = {
            '小分子': 0,
            '聚合物': 0,
            '废弃': 0
        }

        for i, img_data in enumerate(image_data, 1):
            if i % 50 == 0 or i == len(image_data):
                logger.info(f"  分类进度: {i}/{len(image_data)}")

            image_path = img_data['path']
            page_number = img_data['page_number']

            result = self.classifier.classify_image(str(image_path))
            pred_class = result['class']
            classification_stats[pred_class] = classification_stats.get(pred_class, 0) + 1

            # 根据分类结果复制图片到对应文件夹
            type_counters[pred_class] += 1

            if pred_class == '小分子':
                new_filename = f"small_{type_counters['小分子']}.png"
                new_path = small_molecule_dir / new_filename
            elif pred_class == '聚合物':
                new_filename = f"polymer_{type_counters['聚合物']}.png"
                new_path = polymer_dir / new_filename
            else:  # 废弃
                new_filename = f"discarded_{type_counters['废弃']}.png"
                new_path = discarded_dir / new_filename

            # 复制文件到对应文件夹
            import shutil
            shutil.copy2(str(image_path), str(new_path))

            # 保留小分子和聚合物的信息用于识别
            if pred_class in ['小分子', '聚合物']:
                valid_images.append({
                    'path': str(new_path),  # 使用新路径
                    'class': pred_class,
                    'confidence': result['confidence'],
                    'probabilities': result['probabilities'],
                    'page_number': page_number  # 保存页码信息
                })

        logger.info(f"\n分类结果:")
        logger.info(f"  小分子: {classification_stats['小分子']} 张 -> {small_molecule_dir}")
        logger.info(f"  聚合物: {classification_stats['聚合物']} 张 -> {polymer_dir}")
        logger.info(f"  废弃: {classification_stats['废弃']} 张 -> {discarded_dir}")
        logger.info(f"\n有效图片: {len(valid_images)} 张（小分子 + 聚合物）")

        # 清理原始临时图片（已分类保存到对应文件夹）
        logger.info("\n清理临时文件...")
        try:
            import shutil
            shutil.rmtree(temp_output)  # 删除临时文件夹，不影响分类文件夹
            logger.info(f"✅ 已删除临时文件夹: {temp_output}")
        except Exception as e:
            logger.warning(f"⚠️ 删除临时文件夹失败: {e}")

        # ==================== 步骤3: 分离小分子和聚合物 ====================
        logger.info("\n【步骤3】分离小分子和聚合物...")

        small_molecule_images = [img for img in valid_images if img['class'] == '小分子']
        polymer_images = [img for img in valid_images if img['class'] == '聚合物']

        logger.info(f"小分子: {len(small_molecule_images)} 张")
        logger.info(f"聚合物: {len(polymer_images)} 张")

        all_results = []

        # ==================== 步骤4: 识别小分子 ====================
        if small_molecule_images:
            logger.info(f"\n【步骤4】识别小分子 ({len(small_molecule_images)} 张)...")

            for i, img_info in enumerate(small_molecule_images, 1):
                if i % 10 == 0 or i == len(small_molecule_images):
                    logger.info(f"  识别进度: {i}/{len(small_molecule_images)}")

                result = self.recognizer.recognize(img_info['path'])
                result['image_path'] = img_info['path']
                result['classification'] = '小分子'
                result['classification_confidence'] = img_info['confidence']
                result['page_number'] = img_info.get('page_number', '未知')
                result['yolo_processed'] = False  # 标记未经过YOLO处理
                result['original_smiles'] = ''  # 小分子没有Ce替换

                all_results.append(result)

            success_count = sum(1 for r in all_results if r.get('success'))
            logger.info(f"✅ 小分子识别完成: {success_count}/{len(small_molecule_images)} 成功")

        # ==================== 步骤5: YOLO处理聚合物 ====================
        polymer_results = []
        if polymer_images:
            logger.info(f"\n【步骤5】YOLO处理聚合物 ({len(polymer_images)} 张)...")

            # 加载YOLO处理器
            try:
                sys.path.insert(0, str(PROJECT_ROOT / "yolo"))
                from yolo_processor import YOLOProcessor

                yolo_model_path = PROJECT_ROOT / "yolo" / "detect" / "runs" / "obb" / "train6" / "weights" / "best.pt"

                if not yolo_model_path.exists():
                    logger.warning(f"YOLO模型不存在: {yolo_model_path}")
                    logger.warning("聚合物将直接识别，不经过YOLO处理")

                    # 回退：直接识别聚合物
                    for i, img_info in enumerate(polymer_images, 1):
                        if i % 10 == 0 or i == len(polymer_images):
                            logger.info(f"  识别进度: {i}/{len(polymer_images)}")

                        result = self.recognizer.recognize(img_info['path'])
                        result['image_path'] = img_info['path']
                        result['classification'] = '聚合物'
                        result['classification_confidence'] = img_info['confidence']
                        result['page_number'] = img_info.get('page_number', '未知')
                        result['yolo_processed'] = False
                        result['original_smiles'] = ''  # 未经过YOLO处理，无Ce替换

                        polymer_results.append(result)

                else:
                    logger.info(f"加载YOLO模型: {yolo_model_path.name}")
                    yolo_processor = YOLOProcessor(
                        model_path=str(yolo_model_path),
                        use_smart_placement=True
                    )
                    logger.info("✅ YOLO处理器加载成功")

                    # 创建Ce填充图片保存目录
                    ce_filled_dir = output_dir / "images" / "Ce填充"
                    ce_filled_dir.mkdir(parents=True, exist_ok=True)

                    # 处理每张聚合物图片
                    for i, img_info in enumerate(polymer_images, 1):
                        if i % 10 == 0 or i == len(polymer_images):
                            logger.info(f"  处理进度: {i}/{len(polymer_images)}")

                        try:
                            # YOLO检测和Ce填充
                            image = cv2.imread(img_info['path'])
                            detections = yolo_processor.detect_structures(
                                image_path=img_info['path'],
                                conf=0.25,
                                iou=0.5,
                                imgsz=512
                            )

                            if len(detections) > 0:
                                # 不扩展检测框（扩展比例0%）
                                expansion_factor = 1.00
                                expanded_detections = []
                                for det in detections:
                                    expanded_det = det.copy()
                                    # 不扩展，直接使用原始检测框
                                    expanded_detections.append(expanded_det)

                                # Ce填充
                                filled_image = yolo_processor.fill_structures_with_ce(
                                    image=image,
                                    detections=expanded_detections,
                                    atom_text="Ce",
                                    offset_along=25.0,
                                    offset_perpendicular=-15.0
                                )

                                # 保存Ce填充后的图片
                                ce_filled_path = ce_filled_dir / f"{Path(img_info['path']).stem}_ce_filled.png"
                                cv2.imwrite(str(ce_filled_path), filled_image)

                                logger.info(f"    YOLO处理完成，保存到: {ce_filled_path.name}")

                                # 识别Ce填充后的图片
                                result = self.recognizer.recognize(str(ce_filled_path))
                                result['image_path'] = img_info['path']  # 原始图片路径
                                result['ce_filled_image'] = str(ce_filled_path)  # Ce填充图片路径
                                result['classification'] = '聚合物'
                                result['classification_confidence'] = img_info['confidence']
                                result['page_number'] = img_info.get('page_number', '未知')
                                result['yolo_processed'] = True  # 标记经过YOLO处理
                                result['detections'] = len(detections)

                                # Ce/Ge替换为异丙烯基
                                if result.get('success') and result.get('smiles'):
                                    smiles = result['smiles']
                                    has_ce_or_ge = 'Ce' in smiles or '[Ge]' in smiles

                                    if has_ce_or_ge:
                                        original_smiles = smiles
                                        replaced_smiles = SMILESPostProcessor.replace_ce_with_allyl(original_smiles)
                                        result['original_smiles'] = original_smiles  # 保存原始SMILES（含Ce/Ge）
                                        result['smiles'] = replaced_smiles  # 更新为替换后的SMILES
                                        logger.info(f"    Ce/Ge替换: {original_smiles} → {replaced_smiles}")
                                    else:
                                        result['original_smiles'] = ''  # 没有Ce/Ge
                                else:
                                    result['original_smiles'] = ''

                            else:
                                # 未检测到结构，直接识别原图
                                logger.warning(f"    未检测到化学结构，直接识别原图")
                                result = self.recognizer.recognize(img_info['path'])
                                result['image_path'] = img_info['path']
                                result['classification'] = '聚合物'
                                result['classification_confidence'] = img_info['confidence']
                                result['page_number'] = img_info.get('page_number', '未知')
                                result['yolo_processed'] = False
                                result['original_smiles'] = ''  # 未经过YOLO处理，无Ce替换

                            polymer_results.append(result)

                        except Exception as e:
                            logger.error(f"  YOLO处理失败: {e}")
                            # 失败时回退到直接识别
                            result = self.recognizer.recognize(img_info['path'])
                            result['image_path'] = img_info['path']
                            result['classification'] = '聚合物'
                            result['classification_confidence'] = img_info['confidence']
                            result['page_number'] = img_info.get('page_number', '未知')
                            result['yolo_processed'] = False
                            result['error'] = str(e)
                            result['original_smiles'] = ''  # 未经过YOLO处理，无Ce替换

                            polymer_results.append(result)

            except Exception as e:
                logger.error(f"YOLO处理器加载失败: {e}")
                logger.warning("聚合物将直接识别，不经过YOLO处理")

                # 回退：直接识别聚合物
                for i, img_info in enumerate(polymer_images, 1):
                    if i % 10 == 0 or i == len(polymer_images):
                        logger.info(f"  识别进度: {i}/{len(polymer_images)}")

                    result = self.recognizer.recognize(img_info['path'])
                    result['image_path'] = img_info['path']
                    result['classification'] = '聚合物'
                    result['classification_confidence'] = img_info['confidence']
                    result['page_number'] = img_info.get('page_number', '未知')
                    result['yolo_processed'] = False
                    result['original_smiles'] = ''  # 未经过YOLO处理，无Ce替换

                    polymer_results.append(result)

            all_results.extend(polymer_results)

            success_count = sum(1 for r in polymer_results if r.get('success'))
            logger.info(f"✅ 聚合物处理完成: {success_count}/{len(polymer_images)} 成功")

        # ==================== 步骤6: 生成单一Excel表格 ====================
        logger.info("\n【步骤6】生成Excel表格...")

        excel_path = output_dir / f"{pdf_path.stem}_识别结果.xlsx"
        self._generate_single_excel(all_results, excel_path, pdf_path.name)  # 传入PDF文件名
        logger.info(f"✅ 结果表格: {excel_path}")

        # ==================== 总结 ====================
        logger.info("\n" + "=" * 80)
        logger.info("处理完成")
        logger.info("=" * 80)

        summary = {
            'success': True,
            'pdf': str(pdf_path),
            'output_dir': str(output_dir),
            'recognizer': self.recognizer.recognizer_type,
            'total_images': len(image_data),
            'classification': classification_stats,
            'valid_images': len(valid_images),
            'recognition_success': sum(1 for r in all_results if r.get('success')),
            'recognition_total': len(all_results),
            'ce_replacement': {
                'polymers_with_ce_or_ge': sum(1 for r in all_results if r.get('original_smiles') and ('Ce' in r.get('original_smiles', '') or '[Ge]' in r.get('original_smiles', ''))),
                'rdkit_available': RDKIT_AVAILABLE
            },
            'yolo_processing': {
                'polymer_images': len(polymer_images),
                'yolo_processed': sum(1 for r in all_results if r.get('yolo_processed')),
                'ce_filled_dir': str(output_dir / "images" / "Ce填充") if polymer_images else None
            },
            'excel_file': str(excel_path),
            'image_folders': {
                '小分子': str(small_molecule_dir),
                '聚合物': str(polymer_dir),
                '废弃': str(discarded_dir)
            }
        }

        logger.info(f"识别引擎: {summary['recognizer'].upper()}")
        logger.info(f"提取图片: {summary['total_images']} 张")
        logger.info(f"分类保存:")
        logger.info(f"  小分子: {classification_stats['小分子']} 张")
        logger.info(f"  聚合物: {classification_stats['聚合物']} 张")
        logger.info(f"  废弃: {classification_stats['废弃']} 张")
        logger.info(f"\n识别成功: {summary['recognition_success']}/{summary['recognition_total']} 个")

        # Ce/Ge替换统计
        if summary['ce_replacement']['polymers_with_ce_or_ge'] > 0:
            logger.info(f"\nCe/Ge替换:")
            logger.info(f"  含Ce/Ge的SMILES: {summary['ce_replacement']['polymers_with_ce_or_ge']} 个")
            logger.info(f"  RDKit状态: {'可用' if summary['ce_replacement']['rdkit_available'] else '不可用'}")

        # YOLO处理统计
        if polymer_images:
            logger.info(f"\nYOLO处理:")
            logger.info(f"  聚合物图片: {summary['yolo_processing']['polymer_images']} 张")
            logger.info(f"  成功处理: {summary['yolo_processing']['yolo_processed']} 张")
            logger.info(f"  Ce填充图片: {summary['yolo_processing']['ce_filled_dir']}")

        logger.info(f"\n输出文件:")
        logger.info(f"  Excel: {summary['excel_file']}")
        logger.info(f"  小分子图片: {summary['image_folders']['小分子']}")
        logger.info(f"  聚合物图片: {summary['image_folders']['聚合物']}")
        logger.info(f"  废弃图片: {summary['image_folders']['废弃']}")

        # 保存总结
        import json
        summary_path = output_dir / "processing_summary.json"
        with open(summary_path, 'w', encoding='utf-8') as f:
            json.dump(summary, f, indent=2, ensure_ascii=False)

        logger.info(f"\n✅ 处理总结已保存: {summary_path}")

        return summary

    def _generate_single_excel(self, results: List[Dict], output_path: Path, pdf_name: str):
        """
        生成单一Excel表格（包含所有识别结果，嵌入图片）

        Args:
            results: 识别结果列表
            output_path: 输出路径
            pdf_name: PDF文件名（原文献名称）
        """
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Font
        from PIL import Image

        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "识别结果"

        # 设置标题行 - 添加原始SMILES列（显示Ce/Ge替换前后的对比）
        headers = ['序号', '分子结构图', 'SMILES', '原始SMILES(含Ce/Ge)', '分类', 'YOLO处理', '页码', '原文献名称']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)  # 使用Font对象设置粗体

        # 设置列宽
        ws.column_dimensions['A'].width = 8   # 序号
        ws.column_dimensions['B'].width = 25  # 图片
        ws.column_dimensions['C'].width = 50  # SMILES
        ws.column_dimensions['D'].width = 30  # 原始SMILES(含Ce/Ge)
        ws.column_dimensions['E'].width = 10  # 分类
        ws.column_dimensions['F'].width = 10  # YOLO处理
        ws.column_dimensions['G'].width = 8   # 页码
        ws.column_dimensions['H'].width = 40  # 原文献名称

        # 添加数据行
        temp_img_paths = []  # 记录临时文件路径，最后删除

        for i, result in enumerate(results, 2):  # 从第2行开始
            # 序号
            ws.cell(row=i, column=1, value=i-1)

            # 嵌入图片（第2列）
            image_path = result.get('image_path', '')
            if image_path and Path(image_path).exists():
                try:
                    # 加载并调整图片大小
                    pil_img = Image.open(image_path)
                    # 缩放到合适大小（保持比例）
                    max_height = 150  # 增加高度以提高清晰度
                    if pil_img.height > max_height:
                        ratio = max_height / pil_img.height
                        new_width = int(pil_img.width * ratio)
                        pil_img = pil_img.resize((new_width, max_height), Image.Resampling.LANCZOS)

                    # 如果图片太小，适当放大
                    elif pil_img.height < 80:
                        ratio = 80 / pil_img.height
                        new_width = int(pil_img.width * ratio)
                        pil_img = pil_img.resize((new_width, 80), Image.Resampling.LANCZOS)

                    # 保存临时缩放图片（PNG无损格式）
                    temp_img_path = output_path.parent / f"temp_img_{i}.png"
                    pil_img.save(temp_img_path, format='PNG', optimize=False)
                    temp_img_paths.append(temp_img_path)  # 记录临时文件

                    # 插入到Excel（第2列）
                    xl_img = XLImage(temp_img_path)
                    ws.add_image(xl_img, f'B{i}')

                    # 设置行高以适应图片（增加到112像素）
                    ws.row_dimensions[i].height = 112

                except Exception as e:
                    logger.warning(f"  图片嵌入失败 {image_path}: {e}")
                    ws.cell(row=i, column=2, value="图片加载失败")

            # SMILES（第3列）
            ws.cell(row=i, column=3, value=result.get('smiles', '') if result.get('success') else '')

            # 原始SMILES(含Ce/Ge)（第4列）
            original_smiles = result.get('original_smiles', '')
            if original_smiles and ('Ce' in original_smiles or '[Ge]' in original_smiles):
                # 高亮显示含Ce/Ge的SMILES
                cell = ws.cell(row=i, column=4, value=original_smiles)
                from openpyxl.styles import PatternFill
                cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # 黄色高亮
            else:
                ws.cell(row=i, column=4, value='')

            # 分类（第5列）
            ws.cell(row=i, column=5, value=result.get('classification', '未知'))

            # YOLO处理（第6列）
            yolo_processed = result.get('yolo_processed', False)
            ws.cell(row=i, column=6, value='是' if yolo_processed else '否')

            # 页码（第7列）
            ws.cell(row=i, column=7, value=result.get('page_number', '未知'))

            # 原文献名称（第8列）
            ws.cell(row=i, column=8, value=pdf_name)

        # 保存Excel
        wb.save(output_path)

        # Excel保存完成后删除临时文件
        for temp_path in temp_img_paths:
            try:
                if temp_path.exists():
                    temp_path.unlink()
            except Exception as e:
                logger.warning(f"  删除临时文件失败: {e}")

        logger.info(f"  生成表格: {len(results)} 条记录")
        logger.info(f"  成功: {sum(1 for r in results if r.get('success'))} 条")
        logger.info(f"  失败: {sum(1 for r in results if not r.get('success'))} 条")


def main():
    """主函数"""
    parser = argparse.ArgumentParser(
        description='简化集成专利处理系统 - 单一流程，统一识别',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
使用示例:
  # 使用alchemist API识别（默认）
  python simplified_integrated_processor.py --pdf 专利.pdf

  # 使用DECIMER Transformer识别
  python simplified_integrated_processor.py --pdf 专利.pdf --recognizer decimer

  # 指定输出目录
  python simplified_integrated_processor.py --pdf 专利.pdf --output D:/output --recognizer decimer
        '''
    )

    parser.add_argument('--pdf', type=str, required=True, help='PDF文件路径')
    parser.add_argument('--output', type=str, help='输出目录（可选）')
    parser.add_argument(
        '--recognizer',
        type=str,
        choices=['alchemist', 'decimer'],
        default='alchemist',
        help='识别引擎: alchemist (在线API) 或 decimer (离线模型)'
    )

    args = parser.parse_args()

    # 创建处理器
    processor = SimplifiedIntegratedProcessor(recognizer_type=args.recognizer)

    # 处理PDF
    result = processor.process_pdf(
        pdf_path=args.pdf,
        output_dir=args.output
    )

    logger.info("\n" + "=" * 80)
    logger.info("处理完成！")
    logger.info("=" * 80)

    if result['success']:
        logger.info(f"\n✅ 成功处理 {result['valid_images']} 张图片")
        logger.info(f"   识别成功: {result['recognition_success']} 个SMILES")
        logger.info(f"   识别引擎: {result['recognizer'].upper()}")
        logger.info(f"\n输出文件:")
        logger.info(f"   {result['excel_file']}")


if __name__ == "__main__":
    main()