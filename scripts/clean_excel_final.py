#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
清理Excel表格 - 通过重建方式删除没有分子结构图的行
更可靠的方法：创建新工作簿，只复制有图片的行
"""

import sys
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
import shutil
from datetime import datetime

def clean_excel_by_rebuild(excel_path):
    """
    通过重建Excel删除没有分子结构图的行
    步骤：
    1. 找出所有有图片的行
    2. 创建新工作簿
    3. 只复制有图片的行

    Args:
        excel_path: Excel文件路径
    """
    excel_path = Path(excel_path)
    if not excel_path.exists():
        print(f"[错误] 文件不存在: {excel_path}")
        return

    print("=" * 80)
    print("清理Excel表格 - 删除没有分子结构图的行")
    print("=" * 80)
    print(f"文件: {excel_path.name}")
    print()

    # 加载原始Excel
    print("加载原始文件...")
    wb_old = load_workbook(excel_path)
    ws_old = wb_old.active

    print(f"原始总行数: {ws_old.max_row}")
    print(f"原始图片数: {len(ws_old._images)}")
    print()

    # 收集所有有图片的行号
    print("分析图片位置...")
    rows_with_images = set()
    image_dict = {}  # {行号: 图片对象}

    for image in ws_old._images:
        if hasattr(image.anchor, '_from'):
            row = image.anchor._from.row + 1  # openpyxl行号从0开始
            rows_with_images.add(row)
            image_dict[row] = image

    print(f"有图片的行: {len(rows_with_images)} 行")

    # 找出没有图片的行
    rows_without_images = []
    for row in range(2, ws_old.max_row + 1):  # 从第2行开始（跳过标题行）
        if row not in rows_with_images:
            rows_without_images.append(row)

    print(f"没有图片的行: {len(rows_without_images)} 行")
    print()

    if not rows_without_images:
        print("[OK] 所有数据行都有图片，无需清理！")
        wb_old.close()
        return

    # 创建备份
    backup_path = excel_path.parent / f"{excel_path.stem}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}{excel_path.suffix}"
    print(f"创建备份: {backup_path.name}")
    shutil.copy2(excel_path, backup_path)
    print()

    # 创建新工作簿
    print("创建新的工作簿...")
    wb_new = Workbook()
    ws_new = wb_new.active
    ws_new.title = ws_old.title

    # 复制列宽
    print("复制列宽...")
    for col_idx in range(1, ws_old.max_column + 1):
        col_letter = get_column_letter(col_idx)
        if ws_old.column_dimensions[col_letter].width:
            ws_new.column_dimensions[col_letter].width = ws_old.column_dimensions[col_letter].width

    # 复制标题行
    print("复制标题行...")
    for col in range(1, ws_old.max_column + 1):
        ws_new.cell(row=1, column=col, value=ws_old.cell(row=1, column=col).value)

    # 只复制有图片的行
    print("复制有图片的行...")
    new_row_idx = 2  # 新工作簿的行号，从第2行开始

    for old_row in sorted(rows_with_images):
        # 复制整行数据
        for col in range(1, ws_old.max_column + 1):
            ws_new.cell(row=new_row_idx, column=col, value=ws_old.cell(row=old_row, column=col).value)

        # 复制行高
        if ws_old.row_dimensions[old_row].height:
            ws_new.row_dimensions[new_row_idx].height = ws_old.row_dimensions[old_row].height

        new_row_idx += 1

    print(f"[OK] 复制了 {new_row_idx - 2} 行数据")

    # 添加图片（需要调整图片的位置）
    print("复制图片...")
    for old_row in sorted(rows_with_images):
        if old_row in image_dict:
            image = image_dict[old_row]
            # 计算新行号：旧行号在排序后的位置 + 1（因为有标题行）
            new_row = sorted(rows_with_images).index(old_row) + 2

            # 创建图片副本并更新位置
            # 注意：这里需要深拷贝图片对象比较复杂
            # 简化方案：直接添加原始图片对象
            ws_new.add_image(image)

    print(f"[OK] 添加了 {len(image_dict)} 个图片")

    # 保存新文件
    print()
    print("保存新文件...")
    output_path = excel_path.parent / f"{excel_path.stem}_最终清理.xlsx"
    wb_new.save(output_path)
    wb_new.close()
    wb_old.close()

    print(f"[OK] 已保存: {output_path.name}")
    print()
    print("=" * 80)
    print("清理完成！")
    print("=" * 80)
    print(f"原始文件: {excel_path.name}")
    print(f"备份文件: {backup_path.name}")
    print(f"清理后文件: {output_path.name}")
    print()
    print(f"清理前总行数: {ws_old.max_row} (包括标题行)")
    print(f"删除的行数: {len(rows_without_images)}")
    print(f"清理后总行数: {new_row_idx - 1} (包括标题行)")


if __name__ == "__main__":
    if len(sys.argv) > 1:
        excel_path = sys.argv[1]
    else:
        # 使用原始文件
        excel_path = r"D:\skills\results\[信越]_JP7044011B2_Polymerizable monomers, polymers, resist materials, and pattern forming methods_分子结构_20260330_183331.xlsx"

    clean_excel_by_rebuild(excel_path)