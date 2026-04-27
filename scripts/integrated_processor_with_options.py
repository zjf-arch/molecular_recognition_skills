#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
集成处理流程：支持选择alchemist或DECIMER识别
使用方式：
  python integrated_processor_with_options.py --pdf 专利.pdf --recognizer alchemist
  python integrated_processor_with_options.py --pdf 专利.pdf --recognizer decimer
"""

import sys
import os
from pathlib import Path
import argparse
import pandas as pd
from datetime import datetime
from typing import List, Dict

PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from scripts.logger import logger


class SMILESRecognizer:
    """SMILES识别器 - 支持多种识别引擎"""

    def __init__(self, recognizer_type: str = 'alchemist'):
        """
        初始化识别器

        Args:
            recognizer_type: 识别类型 ('alchemist' 或 'decimer')
        """
        self.recognizer_type = recognizer_type.lower()

        if self.recognizer_type == 'alchemist':
            logger.info("使用 alchemist API 识别")
            from scripts.alchemist_api import AlchemistAPI
            self.recognizer = AlchemistAPI(
                api_url="https://api-ocsr.alchemist.iresearch.net.cn/ocsr/",
                headers={"Content-Type": "application/json"}
            )

        elif self.recognizer_type == 'decimer':
            logger.info("使用 DECIMER Transformer 识别")
            try:
                # 先添加路径到sys.path
                decimer_path = PROJECT_ROOT / "DECIMER-Image_Transformer"
                if str(decimer_path) not in sys.path:
                    sys.path.insert(0, str(decimer_path))
                from DECIMER import predict_SMILES
                self.predict_SMILES = predict_SMILES  # 保存为实例变量
                logger.info("✅ DECIMER Transformer 加载成功")
            except Exception as e:
                logger.error(f"DECIMER Transformer 加载失败: {e}")
                raise
        else:
            raise ValueError(f"不支持的识别类型: {recognizer_type}，请选择 'alchemist' 或 'decimer'")

    def recognize(self, image_path: str) -> Dict:
        """
        识别分子结构图生成SMILES

        Args:
            image_path: 图片路径

        Returns:
            {
                'success': bool,
                'smiles': str,
                'confidence': float,
                'recognizer': str
            }
        """
        try:
            if self.recognizer_type == 'alchemist':
                # 使用alchemist API
                result = self.recognizer.recognize_molecule(image_path)
                result['recognizer'] = 'alchemist'
                return result

            elif self.recognizer_type == 'decimer':
                # 使用DECIMER Transformer
                smiles = self.predict_SMILES(image_path)

                return {
                    'success': True,
                    'smiles': smiles,
                    'confidence': 1.0,  # DECIMER不提供置信度
                    'recognizer': 'decimer',
                    'image_path': image_path
                }

        except Exception as e:
            logger.error(f"识别失败 ({self.recognizer_type}): {e}")
            return {
                'success': False,
                'error': str(e),
                'recognizer': self.recognizer_type
            }


class IntegratedPatentProcessorWithOptions:
    """集成专利处理器 - 支持选择识别引擎"""

    def __init__(self, recognizer_type: str = 'alchemist'):
        """
        初始化处理器

        Args:
            recognizer_type: 识别类型 ('alchemist' 或 'decimer')
        """
        logger.info("=" * 80)
        logger.info("集成专利处理系统初始化")
        logger.info(f"识别引擎: {recognizer_type.upper()}")
        logger.info("=" * 80)

        # 1. DECIMER提取器
        logger.info("加载 DECIMER Segmentation...")
        from scripts.decimer_processor_with_page_tracking import DECIMERProcessorWithPageTracking
        self.decimer = DECIMERProcessorWithPageTracking()

        # 2. ResNet18分类器
        logger.info("加载 ResNet18 分类器...")
        from scripts.molecule_classifier import MoleculeImageClassifier
        self.classifier = MoleculeImageClassifier()

        # 3. YOLO处理器
        logger.info("加载 YOLO 处理器...")
        try:
            sys.path.insert(0, str(PROJECT_ROOT / "yolo"))
            from yolo_processor import YOLOProcessor
            self.yolo = YOLOProcessor(
                model_path=str(PROJECT_ROOT / "yolo" / "detect" / "yolo11n-obb.pt"),
                use_smart_placement=False
            )
            self.yolo_available = True
            logger.info("✅ YOLO 处理器加载成功")
        except Exception as e:
            logger.warning(f"⚠️ YOLO 处理器加载失败: {e}")
            self.yolo_available = False

        # 4. SMILES识别器（可选alchemist或DECIMER）
        logger.info(f"加载 {recognizer_type.upper()} 识别器...")
        self.recognizer = SMILESRecognizer(recognizer_type)

        # 5. 表格生成器
        from scripts.table_generator import TableGenerator
        self.table_generator = TableGenerator()

        logger.info("✅ 所有组件加载完成\n")

    def process_pdf(
        self,
        pdf_path: str,
        output_dir: str = None
    ) -> Dict:
        """
        处理单个PDF文件

        Args:
            pdf_path: PDF文件路径
            output_dir: 输出目录

        Returns:
            处理结果字典
        """
        pdf_path = Path(pdf_path)
        if not pdf_path.exists():
            raise FileNotFoundError(f"PDF文件不存在: {pdf_path}")

        # 输出目录 - 缩短路径避免Windows路径长度限制
        if output_dir is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            # 截断文件名避免路径过长（最多保留前30字符）
            stem_short = pdf_path.stem[:30] if len(pdf_path.stem) > 30 else pdf_path.stem
            output_dir = PROJECT_ROOT / "output" / f"{stem_short}_{timestamp}"
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        logger.info("=" * 80)
        logger.info(f"处理专利: {pdf_path.name}")
        logger.info("=" * 80)
        logger.info(f"输出目录: {output_dir}")

        # ==================== 步骤1: DECIMER提取图片 ====================
        logger.info("\n【步骤1】DECIMER 提取化学结构...")

        # 设置DECIMER输出目录 - 使用临时文件夹，稍后会删除
        temp_output = output_dir / "temp_images"
        self.decimer.output_dir = temp_output

        structures = self.decimer.extract_structures_from_pdf(
            str(pdf_path),
            expand=True,
            save_images=True
        )

        if not structures:
            logger.warning("DECIMER 未提取到任何结构")
            return {
                'success': False,
                'error': '未提取到化学结构',
                'pdf': str(pdf_path)
            }

        logger.info(f"✅ DECIMER 提取: {len(structures)} 张图片")

        # 提取图片路径和页码信息
        image_data = []
        for s in structures:
            image_data.append({
                'path': s['image_path'],
                'page_number': s.get('page_number', '未知')  # 页码信息
            })

        # ==================== 步骤2: ResNet18分类并分别保存 ====================
        logger.info("\n【步骤2】ResNet18 分类并分别保存...")

        # 创建三个分类文件夹（与临时文件夹分离）
        small_molecule_dir = output_dir / "images" / "小分子"
        polymer_dir = output_dir / "images" / "聚合物"
        discarded_dir = output_dir / "images" / "废弃"

        small_molecule_dir.mkdir(parents=True, exist_ok=True)
        polymer_dir.mkdir(parents=True, exist_ok=True)
        discarded_dir.mkdir(parents=True, exist_ok=True)

        classified_images = {
            '小分子': [],
            '聚合物': [],
            '废弃': []
        }

        # 用于记录每种类型的计数
        type_counters = {
            '小分子': 0,
            '聚合物': 0,
            '废弃': 0
        }

        for i, img_data in enumerate(image_data, 1):
            if i % 50 == 0 or i == len(image_data):
                logger.info(f"  分类进度: {i}/{len(image_data)}")

            image_path = img_data['path']
            page_number = img_data['page_number']

            result = self.classifier.classify_image(str(image_path))
            pred_class = result['class']

            # 根据分类结果复制图片到对应文件夹
            type_counters[pred_class] += 1

            if pred_class == '小分子':
                new_filename = f"small_{type_counters['小分子']}.png"
                new_path = small_molecule_dir / new_filename
            elif pred_class == '聚合物':
                new_filename = f"polymer_{type_counters['聚合物']}.png"
                new_path = polymer_dir / new_filename
            else:  # 废弃
                new_filename = f"discarded_{type_counters['废弃']}.png"
                new_path = discarded_dir / new_filename

            # 复制文件到对应文件夹
            import shutil
            shutil.copy2(str(image_path), str(new_path))

            if pred_class in classified_images:
                classified_images[pred_class].append({
                    'path': str(new_path),  # 使用新路径
                    'confidence': result['confidence'],
                    'probabilities': result['probabilities'],
                    'page_number': page_number  # 保存页码信息
                })

        logger.info(f"\n分类结果:")
        logger.info(f"  小分子: {len(classified_images['小分子'])} 张 -> {small_molecule_dir}")
        logger.info(f"  聚合物: {len(classified_images['聚合物'])} 张 -> {polymer_dir}")
        logger.info(f"  废弃: {len(classified_images['废弃'])} 张 -> {discarded_dir}")

        # 清理原始临时图片（已分类保存到对应文件夹）
        logger.info("\n清理临时文件...")
        try:
            import shutil
            shutil.rmtree(temp_output)  # 删除临时文件夹，不影响分类文件夹
            logger.info(f"✅ 已删除临时文件夹: {temp_output}")
        except Exception as e:
            logger.warning(f"⚠️ 删除临时文件夹失败: {e}")

        # ==================== 步骤3: 处理小分子 ====================
        logger.info("\n【步骤3】处理小分子...")
        small_molecule_results = []
        small_molecule_images = classified_images['小分子']

        if small_molecule_images:
            logger.info(f"识别 {len(small_molecule_images)} 张小分子图片...")

            for i, img_info in enumerate(small_molecule_images, 1):
                if i % 10 == 0 or i == len(small_molecule_images):
                    logger.info(f"  识别进度: {i}/{len(small_molecule_images)}")

                result = self.recognizer.recognize(img_info['path'])
                result['image_path'] = img_info['path']
                result['confidence'] = img_info['confidence']
                result['type'] = '小分子'
                result['page_number'] = img_info.get('page_number', '未知')  # 添加页码
                small_molecule_results.append(result)

            logger.info(f"✅ 小分子识别完成: {len(small_molecule_results)} 个")

        # ==================== 步骤4: 处理聚合物 ====================
        logger.info("\n【步骤4】处理聚合物...")
        polymer_results = []
        polymer_images = classified_images['聚合物']

        if polymer_images:
            logger.info(f"处理 {len(polymer_images)} 张聚合物图片...")

            for i, img_info in enumerate(polymer_images, 1):
                if i % 10 == 0 or i == len(polymer_images):
                    logger.info(f"  处理进度: {i}/{len(polymer_images)}")

                image_path = img_info['path']

                # YOLO处理（如果可用）
                if self.yolo_available:
                    try:
                        structures = self.yolo.extract_structures(
                            image_path=image_path,
                            conf=0.25,
                            padding=10
                        )

                        if structures:
                            for j, structure in enumerate(structures):
                                yolo_image_path = output_dir / "yolo_extracted" / f"{Path(image_path).stem}_struct_{j+1}.png"
                                yolo_image_path.parent.mkdir(parents=True, exist_ok=True)
                                structure['image_pil'].save(yolo_image_path)

                                result = self.recognizer.recognize(str(yolo_image_path))
                                result['original_image'] = image_path
                                result['yolo_image'] = str(yolo_image_path)
                                result['structure_index'] = j + 1
                                result['confidence'] = img_info['confidence']
                                result['type'] = '聚合物'
                                result['page_number'] = img_info.get('page_number', '未知')  # 添加页码
                                polymer_results.append(result)
                        else:
                            result = self.recognizer.recognize(image_path)
                            result['image_path'] = image_path
                            result['confidence'] = img_info['confidence']
                            result['type'] = '聚合物'
                            result['page_number'] = img_info.get('page_number', '未知')  # 添加页码
                            polymer_results.append(result)

                    except Exception as e:
                        logger.warning(f"  YOLO处理失败: {e}")
                        result = self.recognizer.recognize(image_path)
                        result['image_path'] = image_path
                        result['confidence'] = img_info['confidence']
                        result['type'] = '聚合物'
                        result['page_number'] = img_info.get('page_number', '未知')  # 添加页码
                        polymer_results.append(result)
                else:
                    result = self.recognizer.recognize(image_path)
                    result['image_path'] = image_path
                    result['confidence'] = img_info['confidence']
                    result['type'] = '聚合物'
                    result['page_number'] = img_info.get('page_number', '未知')  # 添加页码
                    polymer_results.append(result)

            logger.info(f"✅ 聚合物处理完成: {len(polymer_results)} 个")

        # ==================== 步骤5: 生成Excel表格 ====================
        logger.info("\n【步骤5】生成Excel表格...")

        small_molecule_excel = None
        if small_molecule_results:
            small_molecule_excel = output_dir / f"{pdf_path.stem}_小分子.xlsx"
            self._generate_excel(small_molecule_results, small_molecule_excel, '小分子', pdf_path.name)
            logger.info(f"✅ 小分子表格: {small_molecule_excel}")

        polymer_excel = None
        if polymer_results:
            polymer_excel = output_dir / f"{pdf_path.stem}_聚合物.xlsx"
            self._generate_excel(polymer_results, polymer_excel, '聚合物', pdf_path.name)
            logger.info(f"✅ 聚合物表格: {polymer_excel}")

        # ==================== 总结 ====================
        logger.info("\n" + "=" * 80)
        logger.info("处理完成")
        logger.info("=" * 80)

        summary = {
            'success': True,
            'pdf': str(pdf_path),
            'output_dir': str(output_dir),
            'recognizer': self.recognizer.recognizer_type,
            'total_images': len(image_paths),
            'classification': {
                '小分子': len(small_molecule_images),
                '聚合物': len(polymer_images),
                '废弃': len(classified_images['废弃'])
            },
            'recognition': {
                '小分子': len(small_molecule_results),
                '聚合物': len(polymer_results)
            },
            'excel_files': {
                '小分子': str(small_molecule_excel) if small_molecule_excel else None,
                '聚合物': str(polymer_excel) if polymer_excel else None
            }
        }

        logger.info(f"识别引擎: {summary['recognizer'].upper()}")
        logger.info(f"图片总数: {summary['total_images']}")
        logger.info(f"分类结果:")
        logger.info(f"  小分子: {summary['classification']['小分子']} 张")
        logger.info(f"  聚合物: {summary['classification']['聚合物']} 张")
        logger.info(f"  废弃: {summary['classification']['废弃']} 张（已过滤）")
        logger.info(f"\n识别结果:")
        logger.info(f"  小分子: {summary['recognition']['小分子']} 个SMILES")
        logger.info(f"  聚合物: {summary['recognition']['聚合物']} 个SMILES")
        logger.info(f"\n输出文件:")
        logger.info(f"  小分子Excel: {summary['excel_files']['小分子']}")
        logger.info(f"  聚合物Excel: {summary['excel_files']['聚合物']}")

        # 保存总结
        import json
        summary_path = output_dir / "processing_summary.json"
        with open(summary_path, 'w', encoding='utf-8') as f:
            json.dump(summary, f, indent=2, ensure_ascii=False)

        logger.info(f"\n✅ 处理总结已保存: {summary_path}")

        return summary

    def _generate_excel(self, results: List[Dict], output_path: Path, data_type: str, pdf_name: str):
        """生成Excel表格（嵌入图片）"""
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Font
        from PIL import Image

        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = data_type

        # 设置标题行 - 添加页码列在倒数第二列
        headers = ['序号', '分子结构图', 'SMILES', '页码', '原文献名称']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)  # 使用Font对象设置粗体

        # 设置列宽
        ws.column_dimensions['A'].width = 8   # 序号
        ws.column_dimensions['B'].width = 25  # 图片
        ws.column_dimensions['C'].width = 50  # SMILES
        ws.column_dimensions['D'].width = 8   # 页码
        ws.column_dimensions['E'].width = 40  # 原文献名称

        # 添加数据行（只添加成功的结果）
        row_num = 2
        temp_img_paths = []  # 记录临时文件路径，最后删除
        for result in results:
            if not result.get('success'):
                continue

            # 序号
            ws.cell(row=row_num, column=1, value=row_num-1)

            # 嵌入图片
            image_path = result.get('yolo_image', result.get('image_path', result.get('original_image', '')))
            if image_path and Path(image_path).exists():
                try:
                    # 加载并调整图片大小
                    pil_img = Image.open(image_path)
                    max_height = 150  # 增加高度以提高清晰度
                    if pil_img.height > max_height:
                        ratio = max_height / pil_img.height
                        new_width = int(pil_img.width * ratio)
                        pil_img = pil_img.resize((new_width, max_height), Image.Resampling.LANCZOS)

                    # 如果图片太小，适当放大
                    elif pil_img.height < 80:
                        ratio = 80 / pil_img.height
                        new_width = int(pil_img.width * ratio)
                        pil_img = pil_img.resize((new_width, 80), Image.Resampling.LANCZOS)

                    # 临时保存（PNG无损格式）
                    temp_img_path = output_path.parent / f"temp_img_{row_num}.png"
                    pil_img.save(temp_img_path, format='PNG', optimize=False)
                    temp_img_paths.append(temp_img_path)  # 记录临时文件

                    # 插入Excel
                    xl_img = XLImage(temp_img_path)
                    ws.add_image(xl_img, f'B{row_num}')

                    # 设置行高（增加到112像素）
                    ws.row_dimensions[row_num].height = 112


                except Exception as e:
                    logger.warning(f"  图片嵌入失败: {e}")
                    ws.cell(row=row_num, column=2, value="图片加载失败")

            # SMILES（第3列）
            ws.cell(row=row_num, column=3, value=result.get('smiles', ''))

            # 页码（第4列）
            ws.cell(row=row_num, column=4, value=result.get('page_number', '未知'))

            # 原文献名称（第5列）
            ws.cell(row=row_num, column=5, value=pdf_name)

            row_num += 1

        # 保存Excel
        wb.save(output_path)

        # Excel保存完成后删除临时文件
        for temp_path in temp_img_paths:
            try:
                if temp_path.exists():
                    temp_path.unlink()
            except Exception as e:
                logger.warning(f"  删除临时文件失败: {e}")

        logger.info(f"  生成 {data_type} 表格: {row_num-2} 条记录")


def main():
    """主函数"""
    parser = argparse.ArgumentParser(
        description='集成专利处理系统 - 支持选择识别引擎',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
使用示例:
  # 使用alchemist API识别（默认）
  python integrated_processor_with_options.py --pdf 专利.pdf

  # 使用DECIMER Transformer识别
  python integrated_processor_with_options.py --pdf 专利.pdf --recognizer decimer

  # 指��输出目录
  python integrated_processor_with_options.py --pdf 专利.pdf --output D:/output --recognizer decimer
        '''
    )

    parser.add_argument('--pdf', type=str, required=True, help='PDF文件路径')
    parser.add_argument('--output', type=str, help='输出目录（可选）')
    parser.add_argument(
        '--recognizer',
        type=str,
        choices=['alchemist', 'decimer'],
        default='alchemist',
        help='识别引擎: alchemist (在线API) 或 decimer (离线模型)'
    )

    args = parser.parse_args()

    # 创建处理器
    processor = IntegratedPatentProcessorWithOptions(recognizer_type=args.recognizer)

    # 处理PDF
    result = processor.process_pdf(
        pdf_path=args.pdf,
        output_dir=args.output
    )

    logger.info("\n" + "=" * 80)
    logger.info("处理完成！")
    logger.info("=" * 80)

    if result['success']:
        logger.info(f"\n✅ 成功处理 {result['total_images']} 张图片")
        logger.info(f"   小分子: {result['recognition']['小分子']} 个SMILES")
        logger.info(f"   聚合物: {result['recognition']['聚合物']} 个SMILES")
        logger.info(f"   识别引擎: {result['recognizer'].upper()}")
        logger.info(f"\n输出文件:")
        logger.info(f"   {result['excel_files']['小分子']}")
        logger.info(f"   {result['excel_files']['聚合物']}")


if __name__ == "__main__":
    main()